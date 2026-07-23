import io
import csv
from typing import List, Dict, Optional
from datetime import datetime
import openpyxl
from fpdf import FPDF

from sqlalchemy.orm import Session
from app.repositories.report_repository import ReportRepository

class ReportingService:
    def __init__(self, repo: ReportRepository):
        self.repo = repo

    def _generate_csv(self, data: List[Dict], headers: List[str]) -> bytes:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        return output.getvalue().encode('utf-8')

    def _generate_excel(self, data: List[Dict], headers: List[str]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append([str(row.get(h, '')) for h in headers])
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _generate_pdf(self, data: List[Dict], headers: List[str], title: str) -> bytes:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        pdf.cell(200, 10, txt=title, new_x="LMARGIN", new_y="NEXT", align="C")
        
        pdf.set_font("Helvetica", size=10)
        # Simplistic table rendering
        col_width = 190 / len(headers) if headers else 190
        for h in headers:
            pdf.cell(col_width, 10, str(h), border=1)
        pdf.ln()
        
        for row in data:
            for h in headers:
                pdf.cell(col_width, 10, str(row.get(h, '')), border=1)
            pdf.ln()
            
        return pdf.output()

    def export_orders(self, db: Session, format: str, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None, status: Optional[str] = None, department: Optional[str] = None) -> bytes:
        data = self.repo.get_orders_report(db, start_date, end_date, status, department)
        headers = ["display_id", "status", "total_price", "created_at", "student_name", "department"]
        return self._export_format(data, headers, format, "Orders Report")

    def export_payments(self, db: Session, format: str, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None) -> bytes:
        data = self.repo.get_payments_report(db, start_date, end_date)
        headers = ["id", "entry_type", "amount", "created_at", "description", "display_id"]
        return self._export_format(data, headers, format, "Payments Report")

    def export_expenses(self, db: Session, format: str, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None) -> bytes:
        data = self.repo.get_expenses_report(db, start_date, end_date)
        headers = ["id", "category", "amount", "payment_method", "description", "expense_date"]
        return self._export_format(data, headers, format, "Expenses Report")

    def export_inventory(self, db: Session, format: str) -> bytes:
        data = self.repo.get_inventory_report(db)
        headers = ["name", "category", "current_stock", "unit_cost", "low_stock_threshold"]
        return self._export_format(data, headers, format, "Inventory Report")

    def _export_format(self, data: List[Dict], headers: List[str], format: str, title: str) -> bytes:
        if format == 'csv':
            return self._generate_csv(data, headers)
        elif format == 'excel':
            return self._generate_excel(data, headers)
        elif format == 'pdf':
            return bytes(self._generate_pdf(data, headers, title))
        else:
            raise ValueError(f"Unsupported format: {format}")
